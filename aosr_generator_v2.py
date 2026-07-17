import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os

def run_calculator():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    os.system('cls' if os.name == 'nt' else 'clear')

    print("ВЫБЕРИТЕ РЕЖИМ:")
    print("1 - Из СМЕТЫ (старый режим)")
    print("2 - Только ТЕКСТ (добавление к содержимому)")
    choice = input("Ваш выбор (1 или 2): ")

    try:
        if choice == "2":
            text_to_add = input("\nВведите текст/добавку: ")
            
            # --- НОВЫЙ ВЫБОР ПОЗИЦИИ ---
            print("Куда добавить текст?")
            print("1 - В НАЧАЛО (как префикс)")
            print("2 - В КОНЕЦ (как суффикс)")
            pos_choice = input("Ваш выбор (1 или 2): ") or "2"
            
            print("\nНастройка диапазона в шаблоне:")
            s_row = int(input("НАЧАЛЬНАЯ строка (например, 1): ") or 1)
            e_row = int(input("КОНЕЧНАЯ строка (например, 278): ") or s_row)
            s_col_let = input("В какой СТОЛБЕЦ добавить (например, E): ").upper() or "E"
            
            template_path = filedialog.askopenfilename(title="Выберите шаблон", filetypes=[("Excel", "*.xlsx")])
            if not template_path: return

            wb = openpyxl.load_workbook(template_path)
            sheet = wb.active
            col_idx = openpyxl.utils.column_index_from_string(s_col_let)

            for r in range(s_row, e_row + 1):
                cell = sheet.cell(row=r, column=col_idx)
                
                # Работа с объединенными ячейками
                target_cell = cell
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            target_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            break
                
                old_val = str(target_cell.value) if target_cell.value is not None else ""
                if old_val.startswith('='): continue 

                # --- ЛОГИКА СКЛЕЙКИ ---
                if pos_choice == "1":
                    target_cell.value = f"{text_to_add} {old_val}".strip()
                else:
                    target_cell.value = f"{old_val} {text_to_add}".strip()

            out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Result_Updated.xlsx")
            if out:
                wb.save(out)
                print(f"\nГОТОВО! Текст успешно добавлен.")
                os.startfile(os.path.dirname(out))

    except Exception as e:
        print(f"Ошибка: {e}")
    finally:
        root.destroy()

if __name__ == "__main__":
    run_calculator()
